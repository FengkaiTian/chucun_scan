"""
Academic Job Scanner - 美国/加拿大学术与科研岗位每日扫描

按 job_titles.json 里的职位名称，在 job_sources.json 注册的各招聘平台上
逐源抓取，去重后输出 xlsx（今日新增 / 全部在招 / 已消失 / 源健康度）。

不做相关性评分——xlsx 携带职位描述原文，评分交给下游 AI。

用法:
    python job_scan.py                  正常扫描
    python job_scan.py --check-sources  探测所有源的健康度（第一次务必先跑这个）
    python job_scan.py --only asabe     只跑指定源，调试用
    python job_scan.py --self-test      离线自测，不联网
"""
import os, sys, re, json, time, html, hashlib, sqlite3, argparse, threading, traceback
import urllib.parse as urlparse
import urllib.robotparser as robotparser
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter

try:
    from urllib3.util.retry import Retry
except ImportError:
    Retry = None

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TITLES_FILE  = os.path.join(SCRIPT_DIR, 'job_titles.json')
SOURCES_FILE = os.path.join(SCRIPT_DIR, 'job_sources.json')
CONFIG_FILE  = os.path.join(SCRIPT_DIR, 'job_config.json')
DB_FILE      = os.path.join(SCRIPT_DIR, 'jobs.db')
OUT_DIR      = os.path.join(SCRIPT_DIR, 'job_output')

# 必须是干净的浏览器 UA。早期版本在末尾缀了联系方式，本意是礼貌，
# 但那是个非标准 UA，Cloudflare 一类的防护会直接判成爬虫并回 403。
# 联系方式改用 HTTP 标准的 From 头传递，见 Http.__init__。
DEFAULT_UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
              'AppleWebKit/537.36 (KHTML, like Gecko) '
              'Chrome/124.0.0.0 Safari/537.36')
USER_AGENT = DEFAULT_UA

HTTP_TIMEOUT   = 25
MAX_WORKERS    = 8
PER_HOST_DELAY = 1.2      # 同一域名两次请求之间的最小间隔（秒）
GONE_AFTER     = 3        # 连续多少次「源正常但没再出现」判定为已消失
DESC_LIMIT     = 2000     # xlsx 里描述列的截断长度


# ══════════════════════════════════════════════════════════════════
#  控制台看板
# ══════════════════════════════════════════════════════════════════

try:
    from rich.console import Console
    from rich.table import Table
    from rich.live import Live
    from rich.panel import Panel
    _HAS_RICH = True
except ImportError:
    _HAS_RICH = False

STATUS_STYLE = {
    'OK':      'green',
    'RUNNING': 'yellow',
    'EMPTY':   'yellow',
    'SKIP':    'blue',
    'HTTP':    'red',
    'NOTFEED': 'magenta',
    'DNS':     'red',
    'ERROR':   'red',
    'PENDING': 'dim',
}


class Dashboard:
    """逐源实时状态看板。rich 可用时是刷新表格，否则退化成逐行打印。"""

    def __init__(self, sources, title='岗位扫描'):
        self.title = title
        self.lock = threading.Lock()
        self.rows = {}
        for s in sources:
            self.rows[s['id']] = {
                'name': s.get('name', s['id']), 'type': s.get('type', '?'),
                'status': 'PENDING', 'found': 0, 'kept': 0, 'detail': '',
            }
        self._live = None
        self._console = Console() if _HAS_RICH else None
        self.t0 = time.time()

    # -- 渲染 ------------------------------------------------------
    def _table(self):
        t = Table(title=None, expand=True, header_style='bold')
        t.add_column('源', style='cyan', no_wrap=True, max_width=34)
        t.add_column('类型', style='dim', no_wrap=True, width=12)
        t.add_column('状态', no_wrap=True, width=8)
        t.add_column('抓到', justify='right', width=6)
        t.add_column('命中', justify='right', width=6)
        t.add_column('说明', style='dim', overflow='fold')
        for r in self.rows.values():
            t.add_row(r['name'], r['type'],
                      f"[{STATUS_STYLE.get(r['status'],'white')}]{r['status']}[/]",
                      str(r['found'] or ''), str(r['kept'] or ''), r['detail'][:70])
        done = sum(1 for r in self.rows.values() if r['status'] != 'PENDING' and r['status'] != 'RUNNING')
        kept = sum(r['kept'] for r in self.rows.values())
        sub = (f"进度 {done}/{len(self.rows)} 源 · 命中 {kept} 条 · "
               f"用时 {time.time()-self.t0:.0f}s")
        return Panel(t, title=f'[bold]{self.title}[/]', subtitle=sub)

    def __enter__(self):
        if _HAS_RICH:
            self._live = Live(self._table(), console=self._console,
                              refresh_per_second=4, transient=False)
            self._live.__enter__()
        else:
            print(f'=== {self.title} ===  (装 rich 可获得实时表格: pip install rich)')
        return self

    def __exit__(self, *exc):
        if self._live:
            self._live.update(self._table())
            self._live.__exit__(*exc)
        return False

    def update(self, sid, **kw):
        with self.lock:
            if sid not in self.rows:
                return
            self.rows[sid].update(kw)
            if self._live:
                self._live.update(self._table())
            elif kw.get('status') in ('OK', 'EMPTY', 'SKIP', 'HTTP', 'ERROR'):
                r = self.rows[sid]
                print(f"  [{r['status']:<5}] {r['name'][:38]:<38} "
                      f"抓到 {r['found']:>4} 命中 {r['kept']:>4}  {r['detail'][:50]}")


_MARKUP_RE = re.compile(r'\[/\]|\[/?[a-z][a-z ]*\]')   # 需同时吃掉 [bold green] 和裸 [/]


def say(msg, style=''):
    if _HAS_RICH:
        Console().print(msg, style=style)
    else:
        print(_MARKUP_RE.sub('', str(msg)))


def log_line(msg):
    """追加一行到运行日志。日志失败不影响主流程。"""
    try:
        os.makedirs(OUT_DIR, exist_ok=True)
        stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        clean = _MARKUP_RE.sub('', str(msg))
        with open(os.path.join(OUT_DIR, 'job_scan_log.txt'), 'a', encoding='utf-8') as f:
            f.write(f'[{stamp}] {clean}\n')
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════
#  HTTP
# ══════════════════════════════════════════════════════════════════

class Http:
    """带重试和按域名限速的 requests 封装。"""

    def __init__(self, cfg=None):
        cfg = cfg or {}
        self.s = requests.Session()
        headers = {
            'User-Agent': cfg.get('user_agent') or DEFAULT_UA,
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept': 'application/rss+xml, application/atom+xml, application/xml, '
                      'application/json, text/html;q=0.9, */*;q=0.8',
        }
        # From 是 HTTP 标准里声明联系方式的头，WAF 不会因此拦截
        if cfg.get('contact_email'):
            headers['From'] = cfg['contact_email']
        self.s.headers.update(headers)
        if Retry is not None:
            retry = Retry(total=2, backoff_factor=1.5,
                          status_forcelist=[429, 500, 502, 503, 504],
                          allowed_methods=frozenset(['GET', 'POST']))
            self.s.mount('https://', HTTPAdapter(max_retries=retry, pool_maxsize=MAX_WORKERS * 2))
            self.s.mount('http://', HTTPAdapter(max_retries=retry, pool_maxsize=MAX_WORKERS * 2))
        self._last = {}
        self._lock = threading.Lock()
        self._robots = {}

    def _throttle(self, url):
        """按域名限速。

        sleep 必须在锁外：放在锁内会让等待某个域名的线程堵住所有其他线程，
        8 个 worker 退化成 1 个，整轮扫描（约 350 次请求）从 1 分钟变成 7 分钟。
        """
        host = urlparse.urlparse(url).netloc
        with self._lock:
            prev = self._last.get(host, 0)
            now = time.time()
            wait = max(0.0, PER_HOST_DELAY - (now - prev))
            # 先把「下一次可请求时刻」占住，别的线程据此排队
            self._last[host] = now + wait
        if wait > 0:
            time.sleep(wait)

    def get(self, url, **kw):
        self._throttle(url)
        kw.setdefault('timeout', HTTP_TIMEOUT)
        return self.s.get(url, **kw)

    def post(self, url, **kw):
        self._throttle(url)
        kw.setdefault('timeout', HTTP_TIMEOUT)
        return self.s.post(url, **kw)

    def robots_ok(self, url):
        """只用于 type=html 的源，尊重 robots.txt。取不到 robots.txt 时放行。"""
        p = urlparse.urlparse(url)
        base = f'{p.scheme}://{p.netloc}'
        with self._lock:
            rp = self._robots.get(base)
        if rp is None:
            rp = robotparser.RobotFileParser()
            rp.set_url(base + '/robots.txt')
            try:
                rp.read()
            except Exception:
                rp = 'allow-all'
            with self._lock:
                self._robots[base] = rp
        if rp == 'allow-all':
            return True
        try:
            return rp.can_fetch(USER_AGENT, url)
        except Exception:
            return True


# ══════════════════════════════════════════════════════════════════
#  文本工具
# ══════════════════════════════════════════════════════════════════

_TAG_RE     = re.compile(r'<[^>]+>')
_WS_RE      = re.compile(r'[ \t ]+')
_NL_RE      = re.compile(r'\n{3,}')
_BR_RE      = re.compile(r'<\s*(br|/p|/div|/li|/tr)\s*/?\s*>', re.I)
_SCRIPT_RE  = re.compile(r'<(script|style)[^>]*>.*?</\1>', re.I | re.S)
_PUNCT_RE   = re.compile(r'[^\w\s/&-]')


def strip_html(s):
    if not s:
        return ''
    s = _SCRIPT_RE.sub(' ', str(s))
    s = _BR_RE.sub('\n', s)
    s = _TAG_RE.sub(' ', s)
    s = html.unescape(s)
    s = _WS_RE.sub(' ', s)
    s = _NL_RE.sub('\n\n', s)
    return s.strip()


def norm_title(s):
    """标题归一化：小写、去标点（保留 / & -）、压空格。"""
    s = html.unescape(str(s or '')).lower()
    s = s.replace('–', '-').replace('—', '-')
    s = _PUNCT_RE.sub(' ', s)
    return _WS_RE.sub(' ', s).strip()


TRACKING_PARAMS = {'utm_source', 'utm_medium', 'utm_campaign', 'utm_term',
                   'utm_content', 'trackid', 'src', 'source', 'ref', 'gh_src'}


def canon_url(url):
    """URL 归一化，用于去重：小写域名、去追踪参数、去末尾斜杠。"""
    if not url:
        return ''
    try:
        p = urlparse.urlsplit(url.strip())
        q = [(k, v) for k, v in urlparse.parse_qsl(p.query, keep_blank_values=True)
             if k.lower() not in TRACKING_PARAMS]
        q.sort()
        path = p.path.rstrip('/') or '/'
        return urlparse.urlunsplit((p.scheme.lower() or 'https', p.netloc.lower(),
                                    path, urlparse.urlencode(q), ''))
    except Exception:
        return url.strip()


def job_id_of(job):
    key = canon_url(job.get('url')) or '|'.join(
        [norm_title(job.get('title')), norm_title(job.get('org')), norm_title(job.get('location'))])
    return hashlib.sha1(key.encode('utf-8', 'replace')).hexdigest()[:16]


US_STATES = {
    'alabama','alaska','arizona','arkansas','california','colorado','connecticut','delaware',
    'florida','georgia','hawaii','idaho','illinois','indiana','iowa','kansas','kentucky',
    'louisiana','maine','maryland','massachusetts','michigan','minnesota','mississippi',
    'missouri','montana','nebraska','nevada','new hampshire','new jersey','new mexico',
    'new york','north carolina','north dakota','ohio','oklahoma','oregon','pennsylvania',
    'rhode island','south carolina','south dakota','tennessee','texas','utah','vermont',
    'virginia','washington','west virginia','wisconsin','wyoming','district of columbia',
}
US_ABBR = {'al','ak','az','ar','ca','co','ct','de','fl','ga','hi','id','il','in','ia','ks',
           'ky','la','me','md','ma','mi','mn','ms','mo','mt','ne','nv','nh','nj','nm','ny',
           'nc','nd','oh','ok','or','pa','ri','sc','sd','tn','tx','ut','vt','va','wa','wv',
           'wi','wy','dc'}
CA_PROV = {'ontario','quebec','québec','british columbia','alberta','manitoba','saskatchewan',
           'nova scotia','new brunswick','newfoundland','prince edward island'}
CA_ABBR = {'on','qc','bc','ab','mb','sk','ns','nb','nl','pe','yt','nt','nu'}

# 明确的境外标记。识别到就返回 'XX'，交给 --country 过滤掉。
# 只列高置信度的国名——识别不出的仍然放行，宁可多留也不误杀。
# 只有这三个是合法的国别取值。源配置里的 BOTH 不属于此列。
VALID_COUNTRIES = {'US', 'CA', 'XX'}

FOREIGN_MARKERS = {
    'united kingdom', 'england', 'scotland', 'wales', 'ireland',
    'netherlands', 'germany', 'france', 'spain', 'italy', 'belgium',
    'switzerland', 'austria', 'denmark', 'sweden', 'norway', 'finland',
    'poland', 'portugal', 'greece', 'czech republic', 'hungary',
    'china', 'japan', 'south korea', 'singapore', 'india', 'israel',
    'australia', 'new zealand', 'brazil', 'chile', 'argentina',
    'mexico', 'south africa', 'saudi arabia', 'united arab emirates', 'qatar',
    'morocco', 'egypt', 'kenya', 'nigeria', 'ethiopia', 'ghana', 'tanzania',
    'turkey', 'iran', 'pakistan', 'bangladesh', 'indonesia', 'malaysia',
    'thailand', 'vietnam', 'philippines', 'taiwan', 'hong kong', 'russia',
    'ukraine', 'romania', 'bulgaria', 'croatia', 'slovenia', 'slovakia',
    'lithuania', 'latvia', 'estonia', 'iceland', 'luxembourg', 'cyprus',
    'colombia', 'peru', 'uruguay', 'ecuador', 'costa rica', 'panama',
    'jordan', 'lebanon', 'oman', 'kuwait', 'bahrain', 'kazakhstan',
}


_SEG_SPLIT = re.compile(r'[,/|;\n·•]+')
_ZIP_RE    = re.compile(r'\b\d[\w-]*\b')


def _clean_seg(s):
    """清洗单个地点片段：小写、去标点、去邮编等数字串、压空格。"""
    s = re.sub(r'[^\w\s-]', ' ', str(s).lower())
    s = _ZIP_RE.sub(' ', s)
    return re.sub(r'\s+', ' ', s).strip()


LOC_HINT_RE = re.compile(
    r'(?:location|based\s+in|located\s+in|position\s+(?:is\s+)?based|duty\s+station|'
    r'place\s+of\s+work|work\s+location)\s*[:\-–]?\s*([^\n.;]{3,80})', re.I)


def country_from_text(text):
    """从正文里找地点线索（Location: / based in / located in …）再判国别。"""
    if not text:
        return ''
    for m in LOC_HINT_RE.finditer(text[:3000]):
        c = guess_country(m.group(1), '')
        if c:
            return c
    return ''


def guess_country(location, default=''):
    """从地点字符串判国别。判不出返回 default。

    调用方绝不应把「来源站点的 country」当 default 传进来：那是「这个源覆盖
    哪些国家」的作用域标记，不是某个岗位的所在地。曾经这么用，导致正文写着
    摩洛哥的岗位、因为挂在标了 US 的源上而被标成美国。

    注意：必须在剥掉逗号之前按分隔符切片，否则 "Gainesville, FL" 会粘成一个
    token 而永远匹配不到州缩写。州/省全称是多词的（south dakota、british
    columbia），所以除整片段外还要比对末尾的 2-3 词 n-gram。
    """
    # 统一在一个地方把 default 规整成合法国别，避免下面任何一条返回路径漏掉
    default = default if default in VALID_COUNTRIES else ''

    raw = str(location or '')
    if not raw.strip():
        return default

    whole = _clean_seg(raw)
    if re.search(r'\bcanada\b', whole):
        return 'CA'
    if re.search(r'\b(usa|united states)\b', whole):
        return 'US'
    for f in FOREIGN_MARKERS:
        if re.search(r'(?<![a-z])' + re.escape(f) + r'(?![a-z])', whole):
            return 'XX'

    cands = []
    for seg in _SEG_SPLIT.split(raw):
        seg = _clean_seg(seg)
        if not seg:
            continue
        cands.append(seg)
        w = seg.split()
        for n in (3, 2, 1):                 # 末尾 n-gram，覆盖 "brookings south dakota"
            if len(w) >= n:
                cands.append(' '.join(w[-n:]))

    # 先比全称（无歧义），再比缩写——'on'/'in'/'or' 这类缩写同时也是英文常用词，
    # 让它们排在全称之后可以避免抢答。
    for c in cands:
        if c in CA_PROV:
            return 'CA'
        if c in US_STATES:
            return 'US'
    for c in cands:
        if c in CA_ABBR:
            return 'CA'
        if c in US_ABBR:
            return 'US'
    return default


# ══════════════════════════════════════════════════════════════════
#  职位名称匹配
# ══════════════════════════════════════════════════════════════════

class Matcher:
    def __init__(self, spec):
        self.settings = spec.get('settings', {})
        self.fuzzy_threshold = self.settings.get('fuzzy_threshold', 88)
        self.require_domain_for_all = self.settings.get('require_domain_for_all', True)
        self.require_domain_for_generic = self.settings.get('require_domain_for_generic', True)
        self.thin_chars = self.settings.get('thin_description_chars', 120)

        self.categories = []
        for c in spec['categories']:
            self.categories.append({
                'id': c['id'], 'label': c['label'], 'generic': c.get('generic', False),
                'regexes': [re.compile(p, re.I) for p in c['patterns']],
                'norm_label': norm_title(c['label']),
            })
        self.negatives = [re.compile(p, re.I) for p in spec.get('negative_title_patterns', [])]
        self.neg_context = [re.compile(p, re.I)
                            for p in spec.get('negative_context_patterns', [])]

        self.domain = {}
        for group, words in spec.get('domain_keywords', {}).items():
            self.domain[group] = [(w, re.compile(r'(?<![a-z])' + re.escape(w) + r'(?![a-z])', re.I))
                                  for w in words]

        self.exclusion = [re.compile(p, re.I) for p in spec.get('exclusion_patterns', [])]
        self.sponsorship = [re.compile(p, re.I) for p in spec.get('sponsorship_patterns', [])]
        self.deadlines = [re.compile(p, re.I) for p in spec.get('deadline_patterns', [])]

        try:
            # 必须用 token_sort_ratio 而非 token_set_ratio：后者对子集给满分，
            # 会把导航链接 "Research" 判成 "Research Assistant Professor"。
            from rapidfuzz import fuzz
            self._fuzz = fuzz.token_sort_ratio
        except ImportError:
            import difflib
            self._fuzz = lambda a, b: difflib.SequenceMatcher(None, a, b).ratio() * 100

    # -- 主入口 ----------------------------------------------------
    def classify(self, title, description='', department='', org='', domain_trusted=False):
        """初筛。返回 (类别名, 命中关键词, 是否需人工确认, 拒绝原因)。

        类别名为 None 表示被筛掉。四层：
          1) 标题负面词      —— 标题里就写明了护理/音乐/社工之类
          2) 标题必须命中目标职位
          3) 上下文负面词    —— 标题看不出，但院系/正文明显不对口
          4) 领域关键词      —— 必须沾边农业/遥感/作物/地理空间
        第 4 层有两条豁免：专业板整源豁免；描述过短时保留但标记需人工确认。
        """
        t = title or ''
        for neg in self.negatives:
            if neg.search(t):
                return None, [], False, 'negative-title'

        nt = norm_title(t)
        if not nt:
            return None, [], False, 'empty-title'

        cat = None
        for c in self.categories:              # 顺序敏感：具体的排在前面
            if any(rx.search(t) for rx in c['regexes']):
                cat = c
                break

        if cat is None:                        # 正则没中，走模糊兜底
            toks = nt.split()
            if len(toks) >= 2:                 # 单词标题不兜底，否则 "Research" 会误判
                best, best_score = None, 0
                for c in self.categories:
                    # 长度护栏：候选明显短于类别名时不比，避免子串式误命中
                    if len(nt) < 0.6 * len(c['norm_label']):
                        continue
                    sc = self._fuzz(nt, c['norm_label'])
                    if sc > best_score:
                        best, best_score = c, sc
                if best is not None and best_score >= self.fuzzy_threshold:
                    cat = best

        if cat is None:
            return None, [], False, 'no-title-match'

        desc = description or ''
        context = '\n'.join([t, department or '', org or '', desc[:600]])
        for neg in self.neg_context:
            if neg.search(context):
                return None, [], False, 'negative-context'

        kws = self.keywords(t + '\n' + desc)
        need_domain = self.require_domain_for_all or (
            cat['generic'] and self.require_domain_for_generic)

        if kws or not need_domain:
            return cat['label'], kws, False, ''

        # 描述近乎为空时判断不了，保留并标记，交给下游细筛。
        thin = len(desc.strip()) < self.thin_chars

        # domain_trusted 只在「描述太短、无从判断」时豁免，不是整站无条件放行。
        # 早先的无条件放行让 Agristok（农业「与生物科学」板）把免疫学、社会学
        # 岗位整批带了进来——描述写得清清楚楚不对口，却因为源被标了 trusted 而通过。
        if thin:
            if domain_trusted or not cat['generic']:
                return cat['label'], kws, True, ''

        return None, kws, False, ('generic-title-without-domain'
                                  if cat['generic'] else 'no-domain-keyword')

    def keywords(self, text):
        if not text:
            return []
        hits = []
        for group, words in self.domain.items():
            for w, rx in words:
                if rx.search(text):
                    hits.append(w)
        # 去掉被更长关键词包含的短词，例如同时命中 "gis" 和 "geospatial" 都保留，
        # 但 "remote sensing" 命中时不必再列 "sensing"
        hits = sorted(set(hits), key=lambda w: (-len(w), w))
        kept = []
        for w in hits:
            if not any(w != k and w in k for k in kept):
                kept.append(w)
        return sorted(kept)

    def work_auth(self, text):
        """判断岗位对「持 OPT、需要担保」的候选人是否可行。

        返回 (结论, 依据原文)。学术岗位大多两者都不写，此时是「未说明」——
        那不代表不能投，高校普遍能办 H-1B/J-1；真正致命的是明确写了要公民、
        不担保、或受出口管制的那些。
        """
        if not text:
            return '未说明', ''
        for rx in self.exclusion:
            mm = rx.search(text)
            if mm:
                return '排除', re.sub(r'\s+', ' ', mm.group(0))[:90]
        for rx in self.sponsorship:
            mm = rx.search(text)
            if mm:
                return '可担保', re.sub(r'\s+', ' ', mm.group(0))[:90]
        return '未说明', ''

    def find_deadline(self, text):
        if not text:
            return ''
        for rx in self.deadlines:
            m = rx.search(text)
            if m:
                return (m.group(1) if m.lastindex else m.group(0)).strip()
        return ''


# ══════════════════════════════════════════════════════════════════
#  抓取器
# ══════════════════════════════════════════════════════════════════

class SourceSkip(Exception):
    """源被有意跳过（例如缺 API key），不算错误。"""


class NotAFeed(Exception):
    """URL 返回了 200，但内容不是 feed。

    最常见的成因是网站改版后旧 RSS 路径 302 到首页——此时 feedparser 解析出
    0 条，若不单独识别就会显示成「返回 0 条」，和「今天确实没有符合的岗位」
    完全无法区分。这个异常就是为了把两者分开。
    """


def _mk(src, title, url, org='', dept='', location='', posted='', desc='', raw=None):
    return {
        'title': strip_html(title), 'url': (url or '').strip(),
        'org': strip_html(org) or src.get('name', ''), 'department': strip_html(dept),
        'location': strip_html(location), 'posted_date': posted or '',
        'description': strip_html(desc), 'source_id': src['id'],
        'source_name': src.get('name', src['id']),
        'country_hint': src.get('country', ''),
        'visa_note': src.get('visa_note', ''), 'raw': raw or {},
    }


def _urls_for(src, cfg):
    """把 {q} 展开成多个 URL。"""
    url = src.get('url', '')
    if '{q}' not in url:
        return [url]
    qs = cfg['query_sets'].get(src.get('q_set', 'narrow'), [])
    return [url.replace('{q}', urlparse.quote_plus(q)) for q in qs] or [url]


def _assert_is_feed(resp, feed):
    """0 条时判断到底是「空 feed」还是「压根不是 feed」。"""
    ctype = (resp.headers.get('Content-Type') or '').lower()
    head = (resp.text or '')[:300].lstrip().lower()
    looks_html = 'html' in ctype or head.startswith(('<!doctype html', '<html'))
    if looks_html or getattr(feed, 'bozo', 0):
        snippet = re.sub(r'\s+', ' ', strip_html((resp.text or '')[:400]))[:110]
        raise NotAFeed(f'非 feed 内容（Content-Type: {ctype or "未声明"}，'
                       f'最终 URL: {resp.url[:70]}）: {snippet}')


def fetch_rss(src, ctx):
    try:
        import feedparser
    except ImportError:
        raise SourceSkip('未安装 feedparser（pip install feedparser）')
    out = []
    for url in _urls_for(src, ctx['cfg']):
        r = ctx['http'].get(url)
        r.raise_for_status()
        feed = feedparser.parse(r.content)
        if not feed.entries:
            _assert_is_feed(r, feed)
            continue
        for e in feed.entries:
            desc = e.get('summary') or e.get('description') or ''
            if e.get('content'):
                desc = e['content'][0].get('value', desc)
            posted = ''
            if e.get('published_parse'):
                posted = time.strftime('%Y-%m-%d', e['published_parse'])
            # 不少招聘 feed 会用自定义元素带地点，feedparser 会原样暴露出来。
            # 取到就能填国别列；取不到则回落到源的 country 字段。
            loc = ''
            for k in ('location', 'job_location', 'joblocation', 'city', 'region'):
                v = e.get(k)
                if isinstance(v, str) and v.strip():
                    loc = v
                    break
            out.append(_mk(src, e.get('title', ''), e.get('link', ''),
                           org=e.get('author', '') or feed.feed.get('title', ''),
                           location=loc, posted=posted, desc=desc, raw=dict(e)))
    return out


def fetch_greenhouse(src, ctx):
    url = f"https://boards-api.greenhouse.io/v1/boards/{src['token']}/jobs?content=true"
    r = ctx['http'].get(url)
    r.raise_for_status()
    out = []
    for j in r.json().get('jobs', []):
        out.append(_mk(src, j.get('title', ''), j.get('absolute_url', ''),
                       org=src.get('name', ''),
                       location=(j.get('location') or {}).get('name', ''),
                       posted=(j.get('updated_at') or '')[:10],
                       desc=j.get('content', ''), raw=j))
    return out


def fetch_lever(src, ctx):
    url = f"https://api.lever.co/v0/postings/{src['token']}?mode=json"
    r = ctx['http'].get(url)
    r.raise_for_status()
    out = []
    for j in r.json():
        cat = j.get('categories') or {}
        posted = ''
        if j.get('createdAt'):
            posted = datetime.fromtimestamp(j['createdAt'] / 1000, timezone.utc).strftime('%Y-%m-%d')
        out.append(_mk(src, j.get('text', ''), j.get('hostedUrl', ''),
                       org=src.get('name', ''), dept=cat.get('team', ''),
                       location=cat.get('location', ''), posted=posted,
                       desc=j.get('descriptionPlain') or j.get('description', ''), raw=j))
    return out


def fetch_ashby(src, ctx):
    url = f"https://api.ashbyhq.com/posting-api/job-board/{src['token']}?includeCompensation=false"
    r = ctx['http'].get(url)
    r.raise_for_status()
    out = []
    for j in r.json().get('jobs', []):
        out.append(_mk(src, j.get('title', ''), j.get('jobUrl', ''),
                       org=src.get('name', ''), dept=j.get('department', ''),
                       location=j.get('location', ''),
                       posted=(j.get('publishedAt') or '')[:10],
                       desc=j.get('descriptionPlain') or j.get('descriptionHtml', ''), raw=j))
    return out


def fetch_smartrecruiters(src, ctx):
    url = f"https://api.smartrecruiters.com/v1/companies/{src['token']}/postings?limit=100"
    r = ctx['http'].get(url)
    r.raise_for_status()
    out = []
    for j in r.json().get('content', []):
        loc = j.get('location') or {}
        out.append(_mk(src, j.get('name', ''),
                       f"https://jobs.smartrecruiters.com/{src['token']}/{j.get('id','')}",
                       org=src.get('name', ''),
                       location=', '.join(filter(None, [loc.get('city'), loc.get('region'), loc.get('country')])),
                       posted=(j.get('releasedDate') or '')[:10],
                       desc=json.dumps(j.get('jobAd', {}), ensure_ascii=False), raw=j))
    return out


def fetch_workable(src, ctx):
    url = f"https://apply.workable.com/api/v1/widget/accounts/{src['token']}?details=true"
    r = ctx['http'].get(url)
    r.raise_for_status()
    out = []
    for j in r.json().get('jobs', []):
        out.append(_mk(src, j.get('title', ''), j.get('url', ''),
                       org=src.get('name', ''), dept=j.get('department', ''),
                       location=', '.join(filter(None, [j.get('city'), j.get('state'), j.get('country')])),
                       posted=(j.get('published_on') or '')[:10],
                       desc=j.get('description', ''), raw=j))
    return out


def fetch_recruitee(src, ctx):
    url = f"https://{src['token']}.recruitee.com/api/offers/"
    r = ctx['http'].get(url)
    r.raise_for_status()
    out = []
    for j in r.json().get('offers', []):
        out.append(_mk(src, j.get('title', ''), j.get('careers_url', ''),
                       org=src.get('name', ''), dept=j.get('department', ''),
                       location=j.get('location', ''), posted=(j.get('published_at') or '')[:10],
                       desc=j.get('description', ''), raw=j))
    return out


def fetch_workday(src, ctx):
    """Workday CxS 接口：POST 分页 JSON。大量高校和公司在用。"""
    host, tenant, site = src['host'], src['tenant'], src['site']
    api = f'https://{host}/wday/cxs/{tenant}/{site}/jobs'
    out, seen = [], set()
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'narrow'), ['']):
        offset = 0
        while offset < 200:                     # 每个关键词最多取 200 条
            body = {'appliedFacets': {}, 'limit': 20, 'offset': offset, 'searchText': q}
            r = ctx['http'].post(api, json=body,
                                 headers={'Accept': 'application/json',
                                          'Content-Type': 'application/json'})
            r.raise_for_status()
            data = r.json()
            posts = data.get('jobPostings', [])
            if not posts:
                break
            for j in posts:
                path = j.get('externalPath', '')
                link = f'https://{host}/{site}{path}' if path else ''
                if link in seen:
                    continue
                seen.add(link)
                out.append(_mk(src, j.get('title', ''), link, org=src.get('name', ''),
                               location=j.get('locationsText', ''),
                               posted=j.get('postedOn', ''),
                               desc=j.get('bulletFields') and ' '.join(map(str, j['bulletFields'])) or '',
                               raw=j))
            if len(posts) < 20:
                break
            offset += 20
    return out


def fetch_peopleadmin(src, ctx):
    """PeopleAdmin 站点通常提供 /postings/search.atom。"""
    try:
        import feedparser
    except ImportError:
        raise SourceSkip('未安装 feedparser')
    base = src['url'].rstrip('/')
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'narrow'), ['']):
        url = f'{base}/postings/search.atom?query={urlparse.quote_plus(q)}'
        r = ctx['http'].get(url)
        r.raise_for_status()
        for e in feedparser.parse(r.content).entries:
            out.append(_mk(src, e.get('title', ''), e.get('link', ''),
                           org=src.get('name', ''),
                           posted=time.strftime('%Y-%m-%d', e['published_parse'])
                                  if e.get('published_parse') else '',
                           desc=e.get('summary', ''), raw=dict(e)))
    return out


def fetch_pageup(src, ctx):
    """PageUp（如 UF explore.jobs）。优先试 JSON 接口，失败则回退 HTML 解析。"""
    base = src['url'].rstrip('/')
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'narrow'), ['']):
        url = f'{base}/en-us/search?keywords={urlparse.quote_plus(q)}'
        r = ctx['http'].get(url, headers={'Accept': 'application/json, text/html'})
        r.raise_for_status()
        ctype = r.headers.get('Content-Type', '')
        if 'json' in ctype:
            for j in (r.json().get('SearchResults') or r.json().get('jobs') or []):
                out.append(_mk(src, j.get('Title') or j.get('title', ''),
                               urlparse.urljoin(base, j.get('Url') or j.get('url', '')),
                               org=src.get('name', ''),
                               location=j.get('Location') or j.get('location', ''),
                               desc=j.get('Summary', ''), raw=j))
        else:
            out.extend(_parse_links(src, r.text, base))
    return out


def fetch_oracle_hcm(src, ctx):
    host = src['host'].rstrip('/')
    site = src.get('site', 'CX')
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'narrow'), ['']):
        params = ('onlyData=true&expand=requisitionList.secondaryLocations'
                  f'&finder=findReqs;siteNumber={site},keyword="{urlparse.quote(q)}",limit=100')
        url = f'https://{host}/hcmRestApi/resources/latest/recruitingCEJobRequisitions?{params}'
        r = ctx['http'].get(url, headers={'Accept': 'application/json'})
        r.raise_for_status()
        for item in r.json().get('items', []):
            for j in item.get('requisitionList', []):
                out.append(_mk(src, j.get('Title', ''),
                               f"https://{host}/{site}/job/{j.get('Id','')}",
                               org=src.get('name', ''),
                               location=j.get('PrimaryLocation', ''),
                               posted=(j.get('PostedDate') or '')[:10],
                               desc=j.get('ShortDescriptionStr', ''), raw=j))
    return out


def fetch_icims(src, ctx):
    base = src['url'].rstrip('/')
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'narrow'), ['']):
        url = f'{base}/jobs/search?searchKeyword={urlparse.quote_plus(q)}&in_iframe=1'
        r = ctx['http'].get(url)
        r.raise_for_status()
        out.extend(_parse_links(src, r.text, base))
    return out


def fetch_adzuna(src, ctx):
    app_id = ctx['cfg']['config'].get('adzuna_app_id', '').strip()
    app_key = ctx['cfg']['config'].get('adzuna_app_key', '').strip()
    if not app_id or not app_key:
        raise SourceSkip('未配置 adzuna_app_id / adzuna_app_key（填入 job_config.json 后生效）')
    tmpl = src['url']
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'core'), []):
        for page in (1, 2):
            url = tmpl.replace('{page}', str(page))
            r = ctx['http'].get(url, params={
                'app_id': app_id, 'app_key': app_key, 'what': q,
                'results_per_page': 50, 'content-type': 'application/json',
            })
            if r.status_code == 429:
                break
            r.raise_for_status()
            results = r.json().get('results', [])
            for j in results:
                out.append(_mk(src, j.get('title', ''), j.get('redirect_url', ''),
                               org=(j.get('company') or {}).get('display_name', ''),
                               location=(j.get('location') or {}).get('display_name', ''),
                               posted=(j.get('created') or '')[:10],
                               desc=j.get('description', ''), raw=j))
            if len(results) < 50:
                break
    return out


def fetch_jsearch(src, ctx):
    """JSearch（RapidAPI / OpenWeb Ninja）。

    这是拿到 LinkedIn / Indeed / Glassdoor / ZipRecruiter 库存的现实路径——
    这四家都已关闭面向个人的公开 API，JSearch 把它们聚合后转售，
    合规责任在服务商。注意它不覆盖 ATS（Workday / Greenhouse / Lever /
    Ashby），那部分由本注册表里的其他源负责，两者互补而非重复。
    """
    key = ctx['cfg']['config'].get('rapidapi_key', '').strip()
    if not key:
        raise SourceSkip('未配置 rapidapi_key —— 这是抓 LinkedIn/Indeed/'
                         'Glassdoor/ZipRecruiter 的唯一途径，填入 job_config.json 后生效')
    host = src.get('rapidapi_host', 'jsearch.p.rapidapi.com')
    country = 'us' if src.get('country') == 'US' else 'ca'
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'core'), []):
        r = ctx['http'].get(f'https://{host}/search',
                            headers={'X-RapidAPI-Key': key, 'X-RapidAPI-Host': host},
                            params={'query': f'{q} in {country}', 'page': '1',
                                    'num_pages': str(src.get('num_pages', 1)),
                                    'country': country, 'date_posted': 'month'})
        r.raise_for_status()
        for j in r.json().get('data') or []:
            loc = ', '.join(filter(None, [j.get('job_city'), j.get('job_state'),
                                          j.get('job_country')]))
            out.append(_mk(src, j.get('job_title', ''),
                           j.get('job_apply_link') or j.get('job_google_link', ''),
                           org=j.get('employer_name', ''),
                           dept=j.get('job_publisher', ''),      # 上游是 LinkedIn 还是 Indeed
                           location=loc,
                           posted=(j.get('job_posted_at_datetime_utc') or '')[:10],
                           desc=j.get('job_description', ''), raw=j))
    return out


def fetch_careerjet(src, ctx):
    """Careerjet 公开 API。免费 key，限速 1000 次/小时，覆盖美加。"""
    affid = ctx['cfg']['config'].get('careerjet_affid', '').strip()
    if not affid:
        raise SourceSkip('未配置 careerjet_affid（免费申请：careerjet.com/partners/api/）')
    locale = 'en_US' if src.get('country') == 'US' else 'en_CA'
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'core'), []):
        r = ctx['http'].get('https://public.api.careerjet.net/search',
                            params={'keywords': q, 'locale_code': locale,
                                    'affid': affid, 'pagesize': 99, 'sort': 'date',
                                    'user_ip': '1.1.1.1', 'user_agent': DEFAULT_UA})
        r.raise_for_status()
        data = r.json()
        if data.get('type') != 'JOBS':
            continue
        for j in data.get('jobs') or []:
            out.append(_mk(src, j.get('title', ''), j.get('url', ''),
                           org=j.get('company', ''), location=j.get('locations', ''),
                           posted=(j.get('date') or '')[:10],
                           desc=j.get('description', ''), raw=j))
    return out


def fetch_jooble(src, ctx):
    """Jooble 公开 API。申请制免费 key，POST JSON。"""
    key = ctx['cfg']['config'].get('jooble_key', '').strip()
    if not key:
        raise SourceSkip('未配置 jooble_key（免费申请：jooble.org/api/about）')
    region = src.get('jooble_region', 'us')
    location = 'United States' if src.get('country') == 'US' else 'Canada'
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'core'), []):
        r = ctx['http'].post(f'https://{region}.jooble.org/api/{key}',
                             json={'keywords': q, 'location': location, 'page': '1'},
                             headers={'Content-Type': 'application/json'})
        r.raise_for_status()
        for j in r.json().get('jobs') or []:
            out.append(_mk(src, j.get('title', ''), j.get('link', ''),
                           org=j.get('company', ''), location=j.get('location', ''),
                           posted=(j.get('updated') or '')[:10],
                           desc=j.get('snippet', ''), raw=j))
    return out


def fetch_usajobs(src, ctx):
    key = ctx['cfg']['config'].get('usajobs_api_key', '').strip()
    email = ctx['cfg']['config'].get('usajobs_email', '').strip()
    if not key or not email:
        raise SourceSkip('未配置 usajobs_api_key / usajobs_email')
    out = []
    for q in ctx['cfg']['query_sets'].get(src.get('q_set', 'core'), []):
        r = ctx['http'].get(src['url'],
                            headers={'Host': 'data.usajobs.gov', 'User-Agent': email,
                                     'Authorization-Key': key},
                            params={'Keyword': q, 'ResultsPerPage': 100})
        r.raise_for_status()
        items = r.json().get('SearchResult', {}).get('SearchResultItems', [])
        for it in items:
            j = it.get('MatchedObjectDescriptor', {})
            locs = '; '.join(l.get('LocationName', '') for l in j.get('PositionLocation', []))
            ud = j.get('UserArea', {}).get('Details', {})
            out.append(_mk(src, j.get('PositionTitle', ''), j.get('PositionURI', ''),
                           org=j.get('OrganizationName', ''),
                           dept=j.get('DepartmentName', ''), location=locs,
                           posted=(j.get('PublicationStartDate') or '')[:10],
                           desc=' '.join(filter(None, [ud.get('JobSummary', ''),
                                                       ud.get('MajorDuties') and
                                                       ' '.join(ud['MajorDuties']) or ''])),
                           raw=j))
    return out


_LINK_RE = re.compile(r'<a\s[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', re.I | re.S)


def _parse_links(src, htmltext, base):
    """从 HTML 里粗提取 <a> 文本作为候选标题。标题匹配阶段会把噪声滤掉。"""
    out, seen = [], set()
    for href, text in _LINK_RE.findall(htmltext or ''):
        title = strip_html(text)
        if not title or len(title) < 6 or len(title) > 200:
            continue
        url = urlparse.urljoin(base, html.unescape(href))
        if url in seen:
            continue
        seen.add(url)
        out.append(_mk(src, title, url, org=src.get('name', '')))
    return out


def fetch_html(src, ctx):
    urls = _urls_for(src, ctx['cfg'])
    out = []
    for url in urls:
        if not ctx['http'].robots_ok(url):
            raise SourceSkip('robots.txt 不允许抓取')
        r = ctx['http'].get(url)
        r.raise_for_status()
        out.extend(_parse_links(src, r.text, url))
    return out


FETCHERS = {
    'rss': fetch_rss, 'greenhouse': fetch_greenhouse, 'lever': fetch_lever,
    'ashby': fetch_ashby, 'smartrecruiters': fetch_smartrecruiters,
    'workable': fetch_workable, 'recruitee': fetch_recruitee,
    'workday': fetch_workday, 'peopleadmin': fetch_peopleadmin,
    'pageup': fetch_pageup, 'oracle_hcm': fetch_oracle_hcm, 'icims': fetch_icims,
    'adzuna': fetch_adzuna, 'usajobs': fetch_usajobs, 'html': fetch_html,
    'jsearch': fetch_jsearch, 'careerjet': fetch_careerjet, 'jooble': fetch_jooble,
}


# ══════════════════════════════════════════════════════════════════
#  存储
# ══════════════════════════════════════════════════════════════════

SCHEMA = """
CREATE TABLE IF NOT EXISTS jobs (
    job_id        TEXT PRIMARY KEY,
    title         TEXT, title_category TEXT, org TEXT, department TEXT,
    location      TEXT, country TEXT, posted_date TEXT, deadline TEXT,
    work_auth     TEXT, work_auth_note TEXT, needs_review TEXT DEFAULT '', url TEXT,
    source_id     TEXT, source_name TEXT, keywords TEXT,
    description   TEXT, raw_json TEXT,
    first_seen    TEXT, last_seen TEXT, missing_runs INTEGER DEFAULT 0,
    status        TEXT DEFAULT 'active'
);
CREATE INDEX IF NOT EXISTS idx_jobs_status ON jobs(status);
CREATE INDEX IF NOT EXISTS idx_jobs_first  ON jobs(first_seen);
CREATE TABLE IF NOT EXISTS runs (
    run_at TEXT PRIMARY KEY, sources_ok INTEGER, sources_bad INTEGER,
    fetched INTEGER, matched INTEGER, new_jobs INTEGER
);
"""


class Store:
    def __init__(self, path=DB_FILE):
        self.db = sqlite3.connect(path)
        self.db.row_factory = sqlite3.Row
        self.db.executescript(SCHEMA)
        # 老库平滑升级：早期版本没有 needs_review 列
        cols = {r[1] for r in self.db.execute('PRAGMA table_info(jobs)')}
        for col in ('needs_review', 'work_auth', 'work_auth_note'):
            if col not in cols:
                self.db.execute(f"ALTER TABLE jobs ADD COLUMN {col} TEXT DEFAULT ''")
        self.db.commit()

    def upsert(self, jobs, today):
        """写入本次抓到的岗位，返回今日新增的 job_id 集合。"""
        new_ids = set()
        cur = self.db.cursor()
        for j in jobs:
            jid = j['job_id']
            row = cur.execute('SELECT job_id FROM jobs WHERE job_id=?', (jid,)).fetchone()
            if row is None:
                new_ids.add(jid)
                cur.execute("""INSERT INTO jobs
                    (job_id,title,title_category,org,department,location,country,posted_date,
                     deadline,work_auth,work_auth_note,needs_review,url,source_id,source_name,
                     keywords,description,raw_json,first_seen,last_seen,missing_runs,status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,'active')""",
                    (jid, j['title'], j['title_category'], j['org'], j['department'],
                     j['location'], j['country'], j['posted_date'], j['deadline'],
                     j['work_auth'], j['work_auth_note'], j['needs_review'], j['url'],
                     j['source_id'], j['source_name'], j['keywords'], j['description'],
                     json.dumps(j.get('raw', {}), ensure_ascii=False, default=str),
                     today, today))
            else:
                cur.execute("""UPDATE jobs SET last_seen=?, missing_runs=0, status='active',
                               deadline=COALESCE(NULLIF(?,''),deadline),
                               description=COALESCE(NULLIF(?,''),description)
                               WHERE job_id=?""",
                            (today, j['deadline'], j['description'], jid))
        self.db.commit()
        return new_ids

    def age_out(self, ok_source_ids, seen_ids, today):
        """只对本次抓取成功的源做老化，避免某个源挂掉就把岗位误判为已消失。"""
        if not ok_source_ids:
            return 0
        qmarks = ','.join('?' * len(ok_source_ids))
        rows = self.db.execute(
            f"SELECT job_id FROM jobs WHERE status='active' AND source_id IN ({qmarks})",
            tuple(ok_source_ids)).fetchall()
        stale = [r['job_id'] for r in rows if r['job_id'] not in seen_ids]
        if not stale:
            return 0
        cur = self.db.cursor()
        cur.executemany('UPDATE jobs SET missing_runs = missing_runs + 1 WHERE job_id=?',
                        [(i,) for i in stale])
        cur.execute(f"UPDATE jobs SET status='gone' WHERE status='active' AND missing_runs>=?",
                    (GONE_AFTER,))
        gone = cur.rowcount
        self.db.commit()
        return gone

    def select(self, where, params=()):
        return [dict(r) for r in self.db.execute(
            f'SELECT * FROM jobs WHERE {where} ORDER BY first_seen DESC, org, title', params)]

    def log_run(self, **kw):
        self.db.execute("""INSERT OR REPLACE INTO runs
            (run_at,sources_ok,sources_bad,fetched,matched,new_jobs) VALUES (?,?,?,?,?,?)""",
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), kw['ok'], kw['bad'],
             kw['fetched'], kw['matched'], kw['new']))
        self.db.commit()


# ══════════════════════════════════════════════════════════════════
#  xlsx 输出
# ══════════════════════════════════════════════════════════════════

# 前两列由 job_rate.py（Gemini）回填；job_scan 自己永远不写它们——
# 本脚本只负责抓取与初筛，相关性判断交给下游模型。
COLUMNS = [
    ('rating', '匹配度', 9), ('open_status', '岗位状态', 10),
    ('job_title', '岗位名称', 46), ('title_category', '类型', 30),
    ('org', '机构', 30), ('department', '院系', 24),
    ('location', '地点', 26), ('country', '国别', 7),
    ('posted_date', '发布日', 12), ('deadline', '截止信息', 22),
    ('work_auth', '身份可行性', 11), ('work_auth_note', '身份依据', 34),
    ('needs_review', '需人工确认', 11),
    ('url', '申请网址', 52), ('source_name', '来源', 26),
    ('matched_keywords', '命中关键词', 40), ('first_seen', '首次发现', 12),
    ('rating_reason', '判定理由', 54),
    ('description', '职位描述', 90),
]


def write_xlsx(path, sheets, source_rows, rejects=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='2F5D3A')
    link_font = Font(color='1F5FA9', underline='single')

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append([c[1] for c in COLUMNS])
        for i, (_, _, width) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
            cell = ws.cell(row=1, column=i)
            cell.font, cell.fill = head_font, head_fill
            cell.alignment = Alignment(vertical='center')
        url_col = [c[0] for c in COLUMNS].index('url') + 1
        for r in rows:
            ws.append([_cell_value(r, key) for key, _, _ in COLUMNS])
            c = ws.cell(row=ws.max_row, column=url_col)
            if c.value and str(c.value).startswith('http'):
                c.hyperlink, c.font = str(c.value), link_font
        ws.freeze_panes = 'A2'
        if ws.max_row >= 1:
            ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{max(ws.max_row,1)}'

    if rejects:
        # 初筛丢弃样本：让你能一眼判断筛选是不是太狠。
        # 最可能误杀的原因排在最前面。
        order = {r: i for i, r in enumerate(REJECT_PRIORITY)}
        ranked = sorted(rejects, key=lambda r: order.get(r['reason'], 99))[:300]
        ws = wb.create_sheet('初筛丢弃样本')
        ws.append(['丢弃原因', '岗位名称', '机构', '地点', '来源', '网址'])
        for i, w in enumerate([30, 46, 28, 24, 26, 52], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
            ws.cell(row=1, column=i).font = head_font
            ws.cell(row=1, column=i).fill = head_fill
        for r in ranked:
            ws.append([REJECT_LABEL.get(r['reason'], r['reason']), r['title'],
                       r['org'], r['location'], r['source_name'], r['url']])
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:F{max(ws.max_row,1)}'

    ws = wb.create_sheet('源健康度')
    ws.append(['源', '类型', '状态', '抓到', '命中', '说明'])
    for i, w in enumerate([36, 14, 10, 8, 8, 80], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=i).font = head_font
        ws.cell(row=1, column=i).fill = head_fill
    for r in source_rows:
        ws.append([r['name'], r['type'], r['status'], r['found'], r['kept'], r['detail']])
    ws.freeze_panes = 'A2'

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


def _cell_value(row, key):
    if key == 'job_title':
        return row.get('title', '')
    if key == 'matched_keywords':
        return row.get('keywords', '')
    v = row.get(key, '')
    if key == 'description':
        v = (v or '')[:DESC_LIMIT]
    return v if v is not None else ''


# ══════════════════════════════════════════════════════════════════
#  主流程
# ══════════════════════════════════════════════════════════════════

def load_json(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def check_deps():
    """启动时自检可选依赖。缺 feedparser 会让 17 个源静默 SKIP，
    看起来像「全都失败」——所以必须在最上方显式告警。"""
    fatal, degraded = [], []
    try:
        __import__('openpyxl')
    except ImportError:
        fatal.append(('openpyxl', '无法生成 xlsx —— 这是脚本的主要产物，缺了跑完也没有输出'))
    try:
        __import__('feedparser')
    except ImportError:
        degraded.append(('feedparser', 'RSS 与 PeopleAdmin 类型的源（共 17 个）全部无法工作'))
    optional = []
    for mod, why in (('rapidfuzz', '标题模糊匹配退化为标准库 difflib，准确率略降'),
                     ('rich',      '状态看板退化为逐行打印')):
        try:
            __import__(mod)
        except ImportError:
            optional.append((mod, why))

    if fatal or degraded:
        say('[bold red]缺少依赖 —— 这是「所有源都失败」最常见的原因[/]')
        for mod, why in fatal + degraded:
            say(f'[red]  {mod:<12} {why}[/]')
        say('[bold]  修复:  pip install -r requirements_jobscan.txt[/]\n')
    if optional:
        for mod, why in optional:
            say(f'[dim]可选依赖 {mod} 未安装：{why}[/]')
        say('')
    if fatal:
        say('[bold red]缺 openpyxl 时不继续 —— 否则会白跑几百次网络请求再报错。[/]')
        return False
    return True


def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            return load_json(CONFIG_FILE)
        except Exception as e:
            say(f'[yellow]job_config.json 解析失败，按空配置继续：{e}[/]')
    return {}


# 每种 HTTP 码对应的修复方向，直接显示在状态表里，省得你再来问我
HTTP_HINT = {
    401: '需要认证 —— 检查 API key',
    403: '被反爬拦截 —— 在 job_config.json 里换个 user_agent 试试，或改用该站的 RSS',
    404: 'URL 不存在 —— 这条 endpoint 是错的，需要修正或关掉',
    410: '接口已下线 —— 把这条 enabled 改成 false',
    429: '限流 —— 调大 PER_HOST_DELAY 或减少 q_set 里的关键词',
    500: '对方服务器出错 —— 过一天再看',
    503: '对方暂时不可用 —— 过一天再看',
}


def probe_url(url, cfg):
    """单独探测一个 URL，打印足够定位问题的信息。"""
    http = Http(cfg)
    say(f'[bold]探测[/] {url}\n')
    say(f'[dim]User-Agent: {http.s.headers.get("User-Agent")}[/]')
    try:
        r = http.get(url, allow_redirects=True)
    except Exception as e:
        say(f'[red]请求失败: {type(e).__name__}: {str(e)[:200]}[/]')
        return 1

    say(f'状态码        {r.status_code}' + (f'  [red]{HTTP_HINT.get(r.status_code,"")}[/]'
                                            if r.status_code >= 400 else ''))
    say(f'Content-Type  {r.headers.get("Content-Type") or "(未声明)"}')
    say(f'最终 URL      {r.url}')
    if r.history:
        say(f'重定向        {len(r.history)} 次 —— 旧路径失效时的典型征兆')
    say(f'响应大小      {len(r.content)} 字节')

    body = (r.text or '')
    head = body[:300].lstrip().lower()
    if head.startswith(('<!doctype html', '<html')):
        say('[magenta]内容是 HTML 网页，不是 feed/JSON[/]')
    try:
        import feedparser
        feed = feedparser.parse(r.content)
        if feed.entries:
            say(f'[green]解析为 feed 成功，{len(feed.entries)} 条[/]')
            for e in feed.entries[:3]:
                say(f'  · {e.get("title","")[:88]}')
        elif not head.startswith(('<!doctype html', '<html')):
            say(f'[yellow]feed 解析出 0 条（bozo={getattr(feed, "bozo", "?")}）[/]')
    except ImportError:
        say('[dim]未装 feedparser，跳过 feed 解析[/]')

    try:
        data = r.json()
        say(f'[green]是合法 JSON，顶层类型 {type(data).__name__}[/]')
        if isinstance(data, dict):
            say(f'  顶层键: {", ".join(list(data)[:12])}')
    except Exception:
        pass

    say('\n[dim]响应开头 400 字：[/]')
    say(re.sub(r'\s+', ' ', strip_html(body[:800]))[:400] or '(空)')
    return 0


def run_source(src, ctx):
    """跑一个源。返回 (状态, 岗位列表, 说明)。绝不抛异常——单源失败不能拖垮整体。"""
    fn = FETCHERS.get(src.get('type'))
    if fn is None:
        return 'ERROR', [], f"未知类型 {src.get('type')}（job_sources.json 写错了）"
    try:
        jobs = fn(src, ctx)
        if jobs:
            return 'OK', jobs, ''
        return 'EMPTY', [], '接口正常，但没有返回任何岗位'
    except SourceSkip as e:
        return 'SKIP', [], str(e)
    except NotAFeed as e:
        return 'NOTFEED', [], str(e)
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else 0
        hint = HTTP_HINT.get(code, '')
        return 'HTTP', [], f'HTTP {code}' + (f' —— {hint}' if hint else '')
    except requests.Timeout:
        return 'ERROR', [], f'超时（>{HTTP_TIMEOUT}s）—— 对方慢或被静默丢包'
    except requests.ConnectionError as e:
        msg = str(e)
        if any(k in msg for k in ('NameResolution', 'Name or service not known',
                                  'nodename nor servname', 'getaddrinfo')):
            return 'DNS', [], '域名解析失败 —— 主机名拼错或该站已不存在'
        if 'ProxyError' in msg or 'Tunnel connection failed' in msg:
            return 'ERROR', [], '代理拒绝 —— 本机/公司网络策略挡住了这个域名'
        if 'SSLError' in msg or 'CertificateError' in msg:
            return 'ERROR', [], 'TLS 证书校验失败 —— 可能是中间人代理'
        return 'ERROR', [], f'连接失败: {msg[:80]}'
    except requests.RequestException as e:
        return 'ERROR', [], f'{type(e).__name__}: {str(e)[:80]}'
    except Exception as e:
        return 'ERROR', [], f'{type(e).__name__}: {str(e)[:80]}'


REJECT_LABEL = {
    'negative-title':   '标题含明确不相关词（护理/音乐/法学等）',
    'negative-context': '院系或正文明确不对口',
    'no-title-match':   '标题不属于目标职位',
    'no-domain-keyword': '未命中任何领域关键词',
    'generic-title-without-domain': '宽泛职位且无领域关键词',
    'empty-title':      '空标题',
    'country':          '不在美国/加拿大',
    'domain-filter':    '--domain-filter 已开且无领域词',
}
# 前两类最可能是误杀，排在丢弃样本表最前面供你复核
REJECT_PRIORITY = ['no-domain-keyword', 'negative-context',
                   'generic-title-without-domain', 'country']


def process(jobs, matcher, keep_countries, domain_filter, trusted_ids=None):
    """初筛 + 字段抽取。返回 (通过的记录, 丢弃明细)。"""
    trusted_ids = trusted_ids or set()
    kept, rejects = [], []

    def drop(j, reason):
        rejects.append({'title': j['title'], 'org': j['org'],
                        'location': j['location'], 'source_name': j['source_name'],
                        'url': j['url'], 'reason': reason})

    for j in jobs:
        cat, kws, needs_review, reason = matcher.classify(
            j['title'], j['description'], j.get('department', ''), j.get('org', ''),
            domain_trusted=j['source_id'] in trusted_ids)
        if cat is None:
            drop(j, reason)
            continue
        if domain_filter and not kws:
            drop(j, 'domain-filter')
            continue
        # 顺序：地点字段 → 正文里的地点线索 → Unknown。
        # 绝不回落到来源站的 country——那会把国外岗位标成美国。
        country = guess_country(j['location'], '') or country_from_text(j['description'])
        if keep_countries and country and country not in keep_countries:
            drop(j, 'country')
            continue
        blob = f"{j['title']}\n{j['description']}"
        auth, auth_note = matcher.work_auth(blob)
        # 源级别的机构性限制（例如佛州 SB 846 对公立校博后的国别限制），
        # 岗位正文里通常一个字都不会提，只能由注册表带进来。
        src_note = j.get('visa_note', '')
        if src_note:
            auth_note = (auth_note + ' | ' if auth_note else '') + src_note
            if auth == '未说明':
                auth = '需注意'
        kept.append({
            'job_id': job_id_of(j), 'title': j['title'], 'title_category': cat,
            'org': j['org'], 'department': j['department'], 'location': j['location'],
            'country': country, 'posted_date': j['posted_date'],
            'deadline': matcher.find_deadline(blob),
            'work_auth': auth, 'work_auth_note': auth_note,
            'needs_review': 'YES' if needs_review else '',
            'url': j['url'], 'source_id': j['source_id'], 'source_name': j['source_name'],
            'keywords': '; '.join(kws), 'description': j['description'],
            'raw': j.get('raw', {}),
        })
    return kept, rejects


def main():
    ap = argparse.ArgumentParser(description='美国/加拿大学术科研岗位每日扫描')
    ap.add_argument('--check-sources', action='store_true', help='只探测各源健康度，不写库不出表')
    ap.add_argument('--only', metavar='ID', help='只跑指定源（可用逗号分隔多个）')
    ap.add_argument('--self-test', action='store_true', help='离线自测，不联网')
    ap.add_argument('--probe', metavar='URL',
                    help='单独探测一个 URL：状态码、Content-Type、重定向、内容类型、前若干字')
    ap.add_argument('--domain-filter', action='store_true',
                    help='要求所有岗位都命中领域关键词（默认只对宽泛职位要求）')
    ap.add_argument('--country', default='US,CA', help='保留的国别，逗号分隔；留空则不过滤')
    ap.add_argument('--workers', type=int, default=MAX_WORKERS)
    args = ap.parse_args()

    if args.self_test:
        return self_test()

    config = load_config()
    if args.probe:
        return probe_url(args.probe, config)

    if not check_deps():
        return 2

    titles = load_json(TITLES_FILE)
    sources_spec = load_json(SOURCES_FILE)
    matcher = Matcher(titles)

    if args.only:
        # 显式点名时忽略 enabled —— 否则刚配好 key、还没改 enabled 的源没法单独测
        wanted = {x.strip() for x in args.only.split(',') if x.strip()}
        sources = [s for s in sources_spec['sources'] if s['id'] in wanted]
        unknown = wanted - {s['id'] for s in sources}
        if unknown:
            say(f'[yellow]job_sources.json 里没有这些 id: {", ".join(sorted(unknown))}[/]')
    else:
        sources = [s for s in sources_spec['sources'] if s.get('enabled', True)]
    if not sources:
        say('[red]没有启用的源。检查 job_sources.json 里的 enabled 字段。[/]')
        return 1

    ctx = {'http': Http(config),
           'cfg': {'query_sets': sources_spec.get('query_sets', {}), 'config': config}}

    keep_countries = {c.strip().upper() for c in args.country.split(',') if c.strip()}
    today = datetime.now().strftime('%Y-%m-%d')
    title = '源健康度探测' if args.check_sources else f'岗位扫描 {today}'

    trusted_ids = {x['id'] for x in sources if x.get('domain_trusted')}
    all_kept, all_rejects, source_rows, ok_ids = [], [], [], []
    fetched_total = 0

    with Dashboard(sources, title) as dash:
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {}
            for s in sources:
                dash.update(s['id'], status='RUNNING')
                futures[pool.submit(run_source, s, ctx)] = s
            for fut in as_completed(futures):
                s = futures[fut]
                status, jobs, detail = fut.result()
                fetched_total += len(jobs)
                kept, rejects = [], []
                if jobs:
                    kept, rejects = process(jobs, matcher, keep_countries,
                                            args.domain_filter, trusted_ids)
                    if not args.check_sources:
                        all_kept.extend(kept)
                        all_rejects.extend(rejects)
                if status == 'OK':
                    ok_ids.append(s['id'])
                    if not detail and rejects:
                        detail = f'初筛丢弃 {len(rejects)} 条'
                dash.update(s['id'], status=status, found=len(jobs),
                            kept=len(kept), detail=detail)
                source_rows.append({'name': s.get('name', s['id']), 'type': s.get('type', ''),
                                    'status': status, 'found': len(jobs), 'kept': len(kept),
                                    'detail': detail})

    ok = sum(1 for r in source_rows if r['status'] == 'OK')
    bad = sum(1 for r in source_rows if r['status'] in ('HTTP', 'ERROR'))

    if args.check_sources:
        from collections import Counter
        tally = Counter(r['status'] for r in source_rows)
        say('\n[bold]探测完成[/]')
        MEANING = {
            'OK':      ('green',   '正常，有岗位返回'),
            'EMPTY':   ('yellow',  '接口通、但当前无岗位 —— 未必是问题'),
            'NOTFEED': ('magenta', 'URL 返回的不是 feed —— endpoint 需要修正'),
            'SKIP':    ('blue',    '主动跳过（多半是缺 API key）'),
            'HTTP':    ('red',     'HTTP 错误 —— 看每行的具体码与建议'),
            'DNS':     ('red',     '域名解析失败 —— 主机名错或站点已不存在'),
            'ERROR':   ('red',     '连接/其他错误'),
        }
        for st, (color, meaning) in MEANING.items():
            if tally.get(st):
                say(f'  [{color}]{st:<8}[/] {tally[st]:>3} 个 —— {meaning}')
        say('\n把 NOTFEED / HTTP 404 / DNS 的条目在 job_sources.json 里修正或改 '
            'enabled=false；HTTP 403 先试换 user_agent。')
        say('[dim]要单独查某个 URL：python job_scan.py --probe "<url>"[/]')
        return 0

    # 同一岗位可能被多个源抓到，先按 job_id 去重（保留描述最长的那条）
    merged = {}
    for j in all_kept:
        prev = merged.get(j['job_id'])
        if prev is None or len(j['description']) > len(prev['description']):
            merged[j['job_id']] = j
    jobs = list(merged.values())

    store = Store()
    new_ids = store.upsert(jobs, today)
    gone = store.age_out(ok_ids, set(merged.keys()), today)

    sheets = {
        '今日新增': store.select('status="active" AND first_seen=?', (today,)),
        '全部在招': store.select('status="active"'),
        '已消失':   store.select('status="gone"'),
    }
    os.makedirs(OUT_DIR, exist_ok=True)
    latest = os.path.join(OUT_DIR, 'job_scan_latest.xlsx')
    dated  = os.path.join(OUT_DIR, f'job_scan_{today}.xlsx')
    write_xlsx(latest, sheets, source_rows, all_rejects)
    write_xlsx(dated, sheets, source_rows, all_rejects)
    store.log_run(ok=ok, bad=bad, fetched=fetched_total, matched=len(jobs), new=len(new_ids))

    review_n = sum(1 for j in jobs if j.get('needs_review'))
    summary = (f'扫描完成  源 {ok} 正常 / {bad} 失败 · 抓取 {fetched_total} 条 · '
               f'初筛丢弃 {len(all_rejects)} 条 · 保留 {len(jobs)} 条'
               f'（其中需人工确认 {review_n}）· 今日新增 {len(new_ids)} · '
               f'新判定已消失 {gone}')
    say('')
    say(f'[bold green]扫描完成[/]  源 {ok} 正常 / {bad} 失败')
    say(f'抓取 {fetched_total} 条 → 初筛丢弃 {len(all_rejects)} 条 → '
        f'保留 {len(jobs)} 条（需人工确认 {review_n}）')
    say(f'[bold]今日新增 {len(new_ids)}[/] · 新判定已消失 {gone}')
    if all_rejects:
        say('[dim]被丢弃的岗位抽样见 xlsx 的「初筛丢弃样本」表——'
            '若发现误杀，调 job_titles.json 的 domain_keywords 即可。[/]')
    say(f'输出: {latest}')
    say(f'      {dated}')
    log_line(summary)
    if bad:
        failed = [r['name'] for r in source_rows if r['status'] in ('HTTP', 'ERROR')]
        say(f'[yellow]有 {bad} 个源失败。跑 --check-sources 看明细，'
            f'然后在 job_sources.json 里修正。[/]')
        log_line('失败的源: ' + ', '.join(failed))
    return 0


# ══════════════════════════════════════════════════════════════════
#  离线自测
# ══════════════════════════════════════════════════════════════════

def self_test():
    say('[bold]离线自测（不联网）[/]\n')
    titles = load_json(TITLES_FILE)
    m = Matcher(titles)
    failures = []

    cases = [
        ('Assistant or Associate Professor in Remote Sensing', 'Assistant/Associate Professor'),
        ('Assistant Professor - Precision Agriculture', 'Assistant Professor'),
        ('Asst. Prof. of Biosystems Engineering', 'Assistant Professor'),
        ('Research Assistant Professor, Plant Phenotyping', 'Research Assistant Professor'),
        ('Assistant Research Professor (Remote Sensing)', 'Assistant Research Professor'),
        ('Teaching Assistant Professor in Agronomy', 'Teaching Assistant Professor'),
        ('Extension Assistant Professor - Precision Ag', 'Extension Assistant Professor'),
        ('Assistant Professor and Extension Specialist, Digital Agriculture',
         'Assistant Professor and Extension Specialist'),
        ('Postdoctoral Research Associate - UAV Phenotyping', 'Postdoctoral Research Associate'),
        ('Post-Doctoral Researcher in Remote Sensing', 'Postdoctoral Researcher'),
        ('Postdoctoral Fellow, Precision Livestock Farming', 'Postdoctoral Fellow'),
        ('Postdoctoral Scholar - Remote Sensing and Data Science', 'Postdoctoral Scholar'),
        ('Presidential Postdoctoral Fellowship in Plant Science',
         'Presidential / Provost Postdoctoral Fellow'),
        ('Senior Research Associate, Crop Modeling', 'Senior Research Associate'),
        ('Research Associate - Hyperspectral Imaging', 'Research Associate'),
        ('Extension Associate, Agronomy', 'Extension Associate'),
        ('Extension Specialist - Precision Agriculture Technologies', 'Extension Specialist'),
        ('Research Data Scientist, Agriculture', 'Research Data Scientist'),
        ('Research Scientist - Remote Sensing', 'Research Scientist'),
        ('Assistant Scientist, Plant Phenotyping', 'Assistant Scientist'),
        ('Associate Scientist - Crop Science', 'Associate Scientist'),
        ('Staff Scientist, Geospatial Analytics', 'Staff Scientist'),
        ('Research Engineer - Agricultural Robotics', 'Research Engineer'),
        ('Computational Scientist, Plant Genomics', 'Computational Scientist'),
        ('Research Program Manager - Digital Agriculture', 'Research Program Manager'),
        ('Phenotyping Facility Manager', 'Phenotyping Facility Manager'),
        ('Core Facility Manager, Imaging', 'Core Facility Manager'),
        ('Field Research Manager - Crop Trials', 'Field Research Manager'),
        ('UAS Program Manager, Agricultural Operations', 'UAS Program Manager'),
        ('Laboratory Manager - Soil and Plant Analysis', 'Laboratory Manager'),
        ('Remote Sensing Specialist', 'Remote Sensing Specialist'),
        ('GIS Specialist, Agriculture and Natural Resources', 'GIS Specialist'),
        ('Research Fellow in Precision Agriculture', 'Research Fellow'),
        ('Lecturer in Agricultural Engineering', 'Lecturer'),
        ('Instructor - Precision Agriculture Technology', 'Instructor'),
        ('Research Specialist - UAV Imagery', 'Research Specialist'),
    ]
    say(f'[bold]1. 职位名称匹配[/]（{len(cases)} 个用例，覆盖全部类别及常见变体）')
    for title, expect in cases:
        got, kws, _rev, reason = m.classify(title, 'remote sensing precision agriculture')
        if got != expect:
            failures.append(f'  标题匹配: {title!r}\n     期望 {expect!r} 实得 {got!r} ({reason})')
    say(f'   {len(cases)-len([f for f in failures if "标题匹配" in f])}/{len(cases)} 通过')

    say('\n[bold]2. 应当被拒绝的标题[/]')
    rejects = [
        ('Clinical Nurse Educator', 'negative'),
        ('Assistant Professor of Medieval History', 'negative'),
        ('Lecturer in Music Theory', 'negative'),
        ('Lecturer', 'generic-no-domain'),
        ('Research Associate', 'generic-no-domain'),
        ('Administrative Assistant', 'no-match'),
    ]
    for title, why in rejects:
        got, _k, _rev, reason = m.classify(title, '')
        if got is not None:
            failures.append(f'  应拒绝但通过了: {title!r} -> {got!r} ({why})')
    say(f'   {len(rejects)-len([f for f in failures if "应拒绝" in f])}/{len(rejects)} 通过')

    say('\n[bold]3. 宽泛职位 + 领域关键词应当通过[/]')
    got, kws, _rev, _r = m.classify('Research Associate',
                                    'Work on UAV hyperspectral imagery of maize canopies.')
    if got != 'Research Associate' or not kws:
        failures.append(f'  宽泛职位带领域词应通过: 实得 {got!r} kws={kws}')
    else:
        say(f'   通过，命中关键词: {"; ".join(kws)}')

    say('\n[bold]3b. 初筛：明显不相关的岗位必须被拦住[/]')
    leaks = [
        ('Assistant Professor',
         'The Department of Music invites applications for a tenure-track position '
         'in jazz performance and improvisation pedagogy.'),
        ('Postdoctoral Scholar',
         'Cancer immunotherapy research in the Department of Oncology, focusing on '
         'CAR-T cell engineering and tumor microenvironment.'),
        ('Research Scientist',
         'Alzheimer disease drug discovery program. Experience with neurodegeneration '
         'models and high-throughput compound screening required.'),
        ('Assistant Professor of Accounting',
         'The College of Business seeks a faculty member to teach auditing and '
         'managerial accounting at the undergraduate and MBA levels.'),
        ('Postdoctoral Research Associate',
         'The Department of French and Italian seeks a postdoctoral associate in '
         'early modern literature and poetics, with language teaching duties.'),
        ('Staff Scientist',
         'Clinical microbiology laboratory supporting patient care and diagnostic '
         'testing. Board certification preferred.'),
        ('Assistant/Associate Professor of Chemistry',
         'Tenure-track position in organic synthesis and catalysis. Candidates must '
         'hold a PhD in Chemistry and establish an externally funded program.'),
        ('Research Engineer',
         'Design and validate automotive powertrain systems including NVH testing, '
         'transmission calibration and vehicle integration for passenger cars.'),
        ('Assistant Professor of Nursing',
         'The School of Nursing seeks faculty to teach in the BSN program and '
         'supervise clinical rotations.'),
        ('Assistant Professor - Social Work',
         'MSW program accreditation and field placement supervision. LCSW required.'),
    ]
    leaked = []
    for title, desc in leaks:
        got, _k, _rev, why = m.classify(title, desc)
        if got is not None:
            leaked.append(f'  初筛泄漏: {title!r} + {desc[:34]!r} -> {got!r}')
    failures.extend(leaked)
    say(f'   {len(leaks)-len(leaked)}/{len(leaks)} 被正确拦住')

    say('\n[bold]3c. 初筛：真实目标岗位必须保留[/]')
    keeps = [
        ('Assistant Professor in Remote Sensing', 'UAV hyperspectral imagery for crops.', False),
        ('Postdoctoral Scholar - Remote Sensing and Data Science', 'machine learning crop health', False),
        ('Assistant Professor of Precision Ag Automation Engineering',
         'agricultural robotics and precision livestock farming', False),
        ('Research Scientist', 'Geospatial data products from satellite remote sensing.', False),
        ('Assistant Professor', 'Biosystems Engineering. Digital agriculture, plant phenotyping.', False),
        ('Postdoctoral Research Associate', '', True),          # RSS 只给标题 -> 保留并标记
        ('Extension Specialist', 'Soil fertility and nutrient management for corn growers.', False),
    ]
    for title, desc, want_review in keeps:
        got, kws, rev, why = m.classify(title, desc)
        if got is None:
            failures.append(f'  初筛误杀: {title!r} ({why})')
        elif rev != want_review:
            failures.append(f'  需人工确认标记错: {title!r} 期望 {want_review} 实得 {rev}')
    say(f'   {len(keeps)} 个用例（含 1 个无描述的、应标记需人工确认）')

    say('\n[bold]3d. 初筛：domain_trusted 源整源豁免[/]')
    got, _k, rev, why = m.classify('Assistant Professor', 'No description available.',
                                   domain_trusted=True)
    if got is None:
        failures.append(f'  domain_trusted 豁免失效: {why}')
    say('   专业板（ASABE / agristok 等）跳过领域词要求')

    say('\n[bold]3e. 模糊匹配不得把导航链接当岗位[/]')
    for junk in ['Research', 'Faculty Research', 'Our Research', 'Search Jobs', 'Contact Us']:
        got, _k, _rev, _w = m.classify(junk, 'agriculture remote sensing')
        if got is not None:
            failures.append(f'  导航链接被误判为岗位: {junk!r} -> {got!r}')
    say('   5 个 HTML 源常见噪声全部拒绝')

    say('\n[bold]4. 字段抽取[/]')
    blob = ('Review of applications will begin October 15, 2026. '
            'U.S. citizenship is required for this position.')
    dl = m.find_deadline(blob)
    if 'October 15, 2026' not in dl:
        failures.append(f'  截止日期抽取失败: {dl!r}')
    blob2 = 'Open until filled. Visa sponsorship is available for exceptional candidates.'
    if 'until filled' not in m.find_deadline(blob2).lower():
        failures.append('  open-until-filled 抽取失败')
    say(f'   截止={dl!r}  滚动={m.find_deadline(blob2)!r}')

    say('\n[bold]4b. 身份可行性（持 OPT 者能不能投，是硬约束）[/]')
    auth_cases = [
        ('U.S. citizenship is required for this position.', '排除'),
        ('Applicants must be authorized to work in the United States without sponsorship.', '排除'),
        ('We are unable to provide visa sponsorship for this role.', '排除'),
        ('This is an export-controlled position; ITAR restrictions apply.', '排除'),
        ('An active security clearance is required.', '排除'),
        ('Visa sponsorship is available for exceptional candidates.', '可担保'),
        ('The university will sponsor H-1B for the successful candidate.', '可担保'),
        ('Seeking a postdoc in UAV remote sensing of crops. Start date flexible.', '未说明'),
    ]
    for text, want in auth_cases:
        got, _note = m.work_auth(text)
        if got != want:
            failures.append(f'  身份判定: {text[:46]!r} 期望 {want!r} 实得 {got!r}')
    say(f'   {len(auth_cases)} 个用例（排除 / 可担保 / 未说明）')

    say('\n[bold]4c. 源级别的机构性限制必须带进表里[/]')
    # 佛州 SB 846 这类限制，岗位正文里一个字都不会提，只能由注册表带进来
    jv = _mk({'id': 'pageup_ufl', 'name': 'UF', 'country': 'US',
              'visa_note': '佛州 SB 846：公立校博后需逐案豁免'},
             'Postdoctoral Associate', 'https://uf.test/1',
             location='Gainesville, FL', desc='UAV remote sensing of crops in Florida.')
    kv, _r = process([jv], m, {'US', 'CA'}, False, set())
    if not kv:
        failures.append('  带 visa_note 的岗位被误丢')
    elif 'SB 846' not in kv[0]['work_auth_note'] or kv[0]['work_auth'] != '需注意':
        failures.append(f'  源级限制未带入: {kv[0]["work_auth"]!r} / {kv[0]["work_auth_note"]!r}')
    else:
        say(f'   UF 岗位 → 身份可行性={kv[0]["work_auth"]!r}，依据带入成功')

    say('\n[bold]5. URL 归一化与去重[/]')
    a = canon_url('https://Example.EDU/jobs/123/?utm_source=x&b=2&a=1')
    b = canon_url('https://example.edu/jobs/123?a=1&b=2')
    if a != b:
        failures.append(f'  URL 归一化不一致:\n     {a}\n     {b}')
    say(f'   {a}')

    say('\n[bold]6. 国别识别[/]')
    for loc, exp in [('Gainesville, FL', 'US'), ('Guelph, ON, Canada', 'CA'),
                     ('Brookings, South Dakota', 'US'), ('London, ON', 'CA'),
                     ('St. Louis, MO 63132', 'US'), ('Vancouver, British Columbia', 'CA'),
                     ('Norwich, United Kingdom', 'XX'), ('Wageningen, Netherlands', 'XX'),
                     ('Zurich, Switzerland', 'XX'), ('Remote', '')]:
        got = guess_country(loc, '')
        if got != exp:
            failures.append(f'  国别识别: {loc!r} 期望 {exp!r} 实得 {got!r}')
    say('   10 个用例（含境外岗位须判为 XX 以便被 --country 过滤）')

    say('\n[bold]6b. 回归：绝不拿来源站国别当岗位国别[/]')
    # 曾经无地点时回落到源的 country，导致正文写着摩洛哥的岗位被标成美国。
    geo_cases = [
        ('US', 'Remote sensing of crops. Position based in Rabat, Morocco.', 'XX', '境外岗位'),
        ('US', 'Remote sensing of crops. Location: Gainesville, FL', 'US', '正文含美国地点'),
        ('US', 'Remote sensing of crops. No location given anywhere here.', '', '无地点须为 Unknown'),
        ('CA', 'UAV phenotyping work. Located in Guelph, Ontario.', 'CA', '正文含加拿大地点'),
    ]
    for src_c, desc, want, label in geo_cases:
        j = _mk({'id': 's', 'name': 'S', 'country': src_c}, 'Postdoctoral Researcher',
                'https://x.test/1', desc=desc)
        got = guess_country(j['location'], '') or country_from_text(j['description'])
        if got != want:
            failures.append(f'  国别推断（{label}）: 期望 {want!r} 实得 {got!r}')
    say(f'   {len(geo_cases)} 个用例（含「源标 US 但正文是摩洛哥」这个真实 bug）')

    say('\n[bold]6c. 回归：domain_trusted 不得整源放行无关领域[/]')
    trust_cases = [
        ('Immunology of T-cell exhaustion in chronic viral infection studied in mouse models.',
         False, '免疫学长描述'),
        ('Department of Sociology seeks a scholar of rural social movements and organizing.',
         False, '社会学长描述'),
        ('UAV hyperspectral imagery for crop nitrogen estimation across many field sites.',
         True, '农业遥感长描述'),
        ('', True, '无描述（专业板常见）'),
    ]
    for desc, want_keep, label in trust_cases:
        j = _mk({'id': 'trusted', 'name': 'T', 'country': 'BOTH'},
                'Postdoctoral Researcher', 'https://x.test/2', desc=desc)
        k, _r = process([j], m, {'US', 'CA'}, False, {'trusted'})
        if bool(k) != want_keep:
            failures.append(f'  domain_trusted（{label}）: 期望 {"保留" if want_keep else "拦住"}，实得相反')
    say(f'   {len(trust_cases)} 个用例（trusted 现在只在描述过短时豁免）')

    # 回归：源配置里的 country 可能是 BOTH，那是「该源覆盖美加」的作用域标记，
    # 不是国别。若把它当国别返回，13 个 BOTH 源（含 ASABE）的岗位会因为
    # 「不在美国/加拿大」被整批静默丢掉，而 RSS 源恰恰从不携带 location。
    if guess_country('', 'BOTH') != '':
        failures.append('  BOTH 被当成国别返回 —— 会导致 13 个源的岗位被整批丢弃')
    both_src = {'id': 'asabe', 'name': 'ASABE', 'country': 'BOTH'}
    probe_jobs = [
        (_mk(both_src, 'Assistant Professor in Remote Sensing', 'https://a.test/1',
             desc='UAV hyperspectral imagery for precision agriculture.'), 1, '无 location 的 BOTH 源'),
        (_mk(both_src, 'Research Scientist', 'https://a.test/2', location='Gainesville, FL',
             desc='remote sensing of crops'), 1, 'BOTH 源 + 美国地点'),
        (_mk(both_src, 'Research Scientist', 'https://a.test/3', location='Norwich, United Kingdom',
             desc='remote sensing of crops'), 0, 'BOTH 源 + 英国地点'),
    ]
    for job, want, label in probe_jobs:
        k, _r = process([job], m, {'US', 'CA'}, False, set())
        if len(k) != want:
            failures.append(f'  国别过滤: {label} 期望保留 {want} 实得 {len(k)}')
    say('   BOTH 源回归用例 3 个（这是曾经会静默吃掉 ASABE 全部岗位的 bug）')

    say('\n[bold]7. 入库、去重与「已消失」老化[/]')
    tmpdb = os.path.join(OUT_DIR, '_selftest.db')
    os.makedirs(OUT_DIR, exist_ok=True)
    if os.path.exists(tmpdb):
        os.remove(tmpdb)
    st = Store(tmpdb)
    fixtures = [
        _mk({'id': 'src_a', 'name': 'Test A'}, 'Assistant Professor of Precision Agriculture',
            'https://a.edu/jobs/1', location='Ames, IA',
            desc='UAV remote sensing. Applications due October 1, 2026.'),
        _mk({'id': 'src_a', 'name': 'Test A'}, 'Postdoctoral Scholar - Remote Sensing',
            'https://a.edu/jobs/2', location='Davis, CA', desc='hyperspectral phenotyping'),
        _mk({'id': 'src_a', 'name': 'Test A'}, 'Staff Nurse', 'https://a.edu/jobs/3'),
    ]
    kept, rejects = process(fixtures, m, {'US', 'CA'}, False)
    if len(kept) != 2 or len(rejects) != 1:
        failures.append(f'  process(): 期望 2 保留/1 丢弃，实得 {len(kept)}/{len(rejects)}')
    if rejects and rejects[0]['reason'] not in REJECT_LABEL:
        failures.append(f'  丢弃原因未登记: {rejects[0]["reason"]!r}')
    d1 = '2026-08-26'
    new1 = st.upsert(kept, d1)
    if len(new1) != 2:
        failures.append(f'  首次入库应有 2 条新增，实得 {len(new1)}')
    new2 = st.upsert(kept, d1)
    if len(new2) != 0:
        failures.append(f'  重复入库不应新增，实得 {len(new2)}')

    kept_one = kept[:1]
    for i in range(GONE_AFTER):
        st.upsert(kept_one, d1)
        st.age_out(['src_a'], {kept_one[0]['job_id']}, d1)
    active = st.select('status="active"')
    goners = st.select('status="gone"')
    if len(active) != 1 or len(goners) != 1:
        failures.append(f'  老化后应为 1 在招/1 消失，实得 {len(active)}/{len(goners)}')

    st.upsert(kept, d1)
    st.age_out([], set(), d1)          # 源全挂时不应误判
    if len(st.select('status="gone"')) != 0:
        failures.append('  源全部失败时不应把岗位判为已消失')
    say(f'   在招 {len(active)} · 已消失 {len(goners)} · 源失败保护 OK')

    say('\n[bold]7a. 限速：sleep 不得在锁内（否则 8 线程退化为 1）[/]')
    import threading as _th
    _h = Http()
    _t0 = time.time()
    _ts = [_th.Thread(target=lambda i=i: _h._throttle(f'https://host{i}.test/x')) for i in range(6)]
    for _t in _ts: _t.start()
    for _t in _ts: _t.join()
    _elapsed = time.time() - _t0
    if _elapsed > 0.5:
        failures.append(f'  6 个不同域名首次请求耗时 {_elapsed:.2f}s —— sleep 可能仍在锁内')
    say(f'   6 个不同域名并发限速 {_elapsed:.2f}s（应接近 0，不同域名互不阻塞）')

    say('\n[bold]7b. 诊断：非 feed 响应必须与「空 feed」区分开[/]')
    class _R:
        def __init__(self, ct, body, url='https://x.test/feed'):
            self.headers = {'Content-Type': ct}; self.text = body; self.url = url
    class _F:
        def __init__(self, bozo=0): self.bozo = bozo; self.entries = []

    # 情况一：网站改版后 RSS 路径重定向到 HTML 首页 —— 必须报 NotAFeed
    try:
        _assert_is_feed(_R('text/html; charset=utf-8',
                           '<!DOCTYPE html><html><body>Page not found</body></html>'), _F(1))
        failures.append('  HTML 冒充 feed 未被识别')
    except NotAFeed as e:
        say(f'   HTML 响应 -> NOTFEED: {str(e)[:66]}…')
    # 情况二：合法但今天没岗位的空 feed —— 不能报错
    try:
        _assert_is_feed(_R('application/rss+xml',
                           '<?xml version="1.0"?><rss><channel></channel></rss>'), _F(0))
        say('   合法空 feed -> 正常放行（不误报）')
    except NotAFeed:
        failures.append('  合法的空 feed 被误判为 NOTFEED')

    say('\n[bold]7c. 诊断：失败类型必须细分[/]')
    def _fake(exc):
        return run_source({'id': 't', 'name': 't', 'type': '_fake'},
                          {'cfg': {}, 'http': None, '_raise': exc})
    FETCHERS['_fake'] = lambda src, ctx: (_ for _ in ()).throw(ctx['_raise'])
    resp404 = requests.Response(); resp404.status_code = 404
    resp403 = requests.Response(); resp403.status_code = 403
    checks = [
        (requests.HTTPError(response=resp404), 'HTTP', '404'),
        (requests.HTTPError(response=resp403), 'HTTP', '403'),
        (requests.ConnectionError('NameResolutionError host'), 'DNS', '解析'),
        (requests.Timeout(), 'ERROR', '超时'),
        (NotAFeed('不是 feed'), 'NOTFEED', 'feed'),
        (SourceSkip('缺 key'), 'SKIP', 'key'),
    ]
    for exc, want_status, want_in in checks:
        status, _j, detail = _fake(exc)
        if status != want_status or want_in not in detail:
            failures.append(f'  失败分类错: {type(exc).__name__} -> {status} / {detail!r}')
    del FETCHERS['_fake']
    say(f'   {len(checks)} 种失败各自归类正确，且带修复建议')

    say('\n[bold]7d. 新增聚合器连接器（覆盖 LinkedIn/Indeed/Glassdoor/ZipRecruiter）[/]')
    for t in ('jsearch', 'careerjet', 'jooble'):
        if t not in FETCHERS:
            failures.append(f'  抓取器未注册: {t}')
    # 无 key 时必须干净地 SKIP，而不是报错
    for t, sid in (('jsearch', 'jsearch_us'), ('careerjet', 'careerjet_us'),
                   ('jooble', 'jooble_us')):
        status, _j, detail = run_source(
            {'id': sid, 'name': sid, 'type': t, 'country': 'US', 'q_set': 'core'},
            {'cfg': {'query_sets': {'core': ['remote sensing']}, 'config': {}},
             'http': None})
        if status != 'SKIP':
            failures.append(f'  {t} 缺 key 时应 SKIP，实得 {status}: {detail}')
    say('   3 个连接器已注册，且缺 key 时干净跳过（不影响其余源）')

    say('\n[bold]8. xlsx 生成[/]')
    out = os.path.join(OUT_DIR, '_selftest.xlsx')
    write_xlsx(out, {'今日新增': st.select('status="active"'),
                     '全部在招': st.select('status="active"'),
                     '已消失': st.select('status="gone"')},
               [{'name': 'Test A', 'type': 'rss', 'status': 'OK',
                 'found': 3, 'kept': 2, 'detail': ''}],
               rejects)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    if set(wb.sheetnames) != {'今日新增', '全部在招', '已消失', '初筛丢弃样本', '源健康度'}:
        failures.append(f'  工作表不齐: {wb.sheetnames}')
    ws = wb['全部在招']
    hdr = [c.value for c in ws[1]]
    if hdr != [c[1] for c in COLUMNS]:
        failures.append('  表头与 COLUMNS 不一致')
    # 「匹配度」列存在，但必须由 job_rate.py 回填。job_scan 自测的数据里
    # 没有评级，所以这一列此刻应当是空的——不空就说明抓取侧偷偷打分了。
    if '匹配度' not in hdr or '岗位状态' not in hdr:
        failures.append('  缺少「匹配度」/「岗位状态」列（job_rate 回填用）')
    elif ws.cell(row=2, column=hdr.index('匹配度') + 1).value:
        failures.append('  job_scan 不应自己填「匹配度」——评级归 job_rate.py')
    desc_idx = hdr.index('职位描述') + 1
    if not ws.cell(row=2, column=desc_idx).value:
        failures.append('  职位描述列为空——下游 AI 将无法判断相关性')
    if '需人工确认' not in hdr:
        failures.append('  缺少「需人工确认」列')
    for c in ('身份可行性', '身份依据'):
        if c not in hdr:
            failures.append(f'  缺少「{c}」列')
    rej_ws = wb['初筛丢弃样本']
    if rej_ws.max_row < 2:
        failures.append('  「初筛丢弃样本」表没有数据行')
    say(f'   工作表 {wb.sheetnames}')
    say(f'   列数 {len(hdr)} · 匹配度列留空待评 · 描述列有内容')

    st.db.close()
    for f in (tmpdb, out):
        if os.path.exists(f):
            os.remove(f)

    say('')
    if failures:
        say(f'[bold red]自测失败 {len(failures)} 项[/]')
        for f in failures:
            say(f'[red]{f}[/]')
        return 1
    say('[bold green]全部自测通过[/]')
    say('[dim]注意：自测只覆盖离线逻辑。各招聘站的接口连通性必须用 --check-sources 在联网环境验证。[/]')
    return 0


if __name__ == '__main__':
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        say('\n[yellow]已中断[/]')
        sys.exit(130)
    except Exception:
        traceback.print_exc()
        sys.exit(1)
